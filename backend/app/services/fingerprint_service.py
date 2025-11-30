import hashlib
import json
import re
from typing import Dict, List, Any
import openpyxl
from pathlib import Path


class FingerprintService:
    """Excel 파일 핑거프린트 생성 서비스"""

    def generate_fingerprint(self, file_path: str) -> str:
        """Excel 파일의 핑거프린트 생성"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        components = {
            "row_count_range": self._get_range_bucket(sheet.max_row),
            "col_count_range": self._get_range_bucket(sheet.max_column),
            "header_pattern": self._extract_header_pattern(sheet),
            "data_type_pattern": self._extract_data_types(sheet),
            "merged_cells_count": len(list(sheet.merged_cells.ranges)),
            "keywords": self._extract_keywords(sheet),
        }

        return hashlib.sha256(
            json.dumps(components, sort_keys=True).encode()
        ).hexdigest()[:16]

    def _get_range_bucket(self, value: int) -> str:
        """값을 범위 버킷으로 변환"""
        if value <= 10:
            return "small"
        elif value <= 50:
            return "medium"
        elif value <= 200:
            return "large"
        else:
            return "xlarge"

    def _extract_header_pattern(self, sheet) -> List[str]:
        """상위 3행의 헤더 패턴 추출"""
        patterns = []
        for row_num in range(1, min(4, sheet.max_row + 1)):
            row_pattern = []
            for cell in sheet[row_num]:
                if cell.value:
                    # 정규화: 숫자는 '#', 텍스트는 첫 글자만
                    value = str(cell.value)
                    if re.match(r'^[\d.,]+$', value):
                        row_pattern.append('#')
                    else:
                        row_pattern.append(value[:3])
                else:
                    row_pattern.append('')
            patterns.append('|'.join(row_pattern[:10]))  # 최대 10열
        return patterns

    def _extract_data_types(self, sheet) -> str:
        """데이터 타입 패턴 추출 (4-10행)"""
        type_grid = []
        for row_num in range(4, min(11, sheet.max_row + 1)):
            row_types = []
            for cell in list(sheet[row_num])[:10]:  # 최대 10열
                if cell.value is None:
                    row_types.append('E')  # Empty
                elif isinstance(cell.value, (int, float)):
                    row_types.append('N')  # Number
                else:
                    row_types.append('T')  # Text
            type_grid.append(''.join(row_types))
        return ','.join(type_grid)

    def _extract_keywords(self, sheet) -> List[str]:
        """주요 키워드 추출"""
        keywords_to_find = [
            '모델', 'model', '품목', '제품',
            '주차', 'week', '일', 'day', '날짜', 'date',
            '수량', 'qty', 'quantity', '생산',
            '합계', 'total', 'sum'
        ]

        found_keywords = []
        for row in sheet.iter_rows(min_row=1, max_row=5, max_col=10):
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value).lower()
                    for keyword in keywords_to_find:
                        if keyword.lower() in cell_text:
                            found_keywords.append(keyword.lower())

        return list(set(found_keywords))

    def calculate_similarity(
        self,
        fingerprint1: str,
        fingerprint2: str,
        file1_path: str = None,
        file2_path: str = None
    ) -> float:
        """두 핑거프린트의 유사도 계산 (0-100%)"""
        # 기본: 핑거프린트 직접 비교
        if fingerprint1 == fingerprint2:
            return 100.0

        # 상세 비교가 필요한 경우 (파일 경로 제공 시)
        if file1_path and file2_path:
            return self._detailed_similarity(file1_path, file2_path)

        # 해시 기반 부분 일치 검사
        common = sum(1 for a, b in zip(fingerprint1, fingerprint2) if a == b)
        return (common / len(fingerprint1)) * 100

    def _detailed_similarity(self, file1_path: str, file2_path: str) -> float:
        """두 파일의 상세 유사도 계산"""
        try:
            import jellyfish

            wb1 = openpyxl.load_workbook(file1_path, data_only=True)
            wb2 = openpyxl.load_workbook(file2_path, data_only=True)
            sheet1, sheet2 = wb1.active, wb2.active

            scores = []

            # 1. 구조 유사도 (40%)
            row_diff = abs(sheet1.max_row - sheet2.max_row)
            col_diff = abs(sheet1.max_column - sheet2.max_column)
            structure_score = max(0, 100 - (row_diff + col_diff) * 5)
            scores.append(structure_score * 0.4)

            # 2. 헤더 유사도 (40%)
            header1 = self._get_header_text(sheet1)
            header2 = self._get_header_text(sheet2)
            header_similarity = jellyfish.jaro_winkler_similarity(header1, header2)
            scores.append(header_similarity * 100 * 0.4)

            # 3. 키워드 유사도 (20%)
            kw1 = set(self._extract_keywords(sheet1))
            kw2 = set(self._extract_keywords(sheet2))
            if kw1 or kw2:
                keyword_score = len(kw1 & kw2) / len(kw1 | kw2) * 100
            else:
                keyword_score = 50
            scores.append(keyword_score * 0.2)

            return sum(scores)
        except Exception:
            return 50.0  # 오류 시 중간값

    def _get_header_text(self, sheet) -> str:
        """헤더 텍스트 추출"""
        texts = []
        for row_num in range(1, min(4, sheet.max_row + 1)):
            for cell in list(sheet[row_num])[:10]:
                if cell.value:
                    texts.append(str(cell.value))
        return ' '.join(texts)


fingerprint_service = FingerprintService()
