import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from PIL import Image as PILImage
import tempfile
import io
import re
from datetime import datetime, timedelta


class ExcelService:
    """Excel 파일 처리 서비스"""

    def read_excel(self, file_path: str) -> pd.DataFrame:
        """Excel 파일 읽기"""
        return pd.read_excel(file_path)

    def get_sheet_info(self, file_path: str) -> Dict[str, Any]:
        """Excel 시트 정보 가져오기"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        return {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "merged_cells": [str(mc) for mc in sheet.merged_cells.ranges],
            "sheet_names": wb.sheetnames
        }

    def excel_to_image(self, file_path: str, output_path: Optional[str] = None) -> str:
        """Excel 시트를 이미지로 변환 (스크린샷 시뮬레이션)"""
        # 실제 구현에서는 xlsx2img 또는 win32com 사용
        # 여기서는 간단한 텍스트 기반 이미지 생성
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # 데이터 읽기
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 50),
                                    min_col=1, max_col=min(sheet.max_column, 20)):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            data.append(row_data)

        # PIL로 이미지 생성
        from PIL import Image, ImageDraw, ImageFont

        cell_width = 100
        cell_height = 25
        padding = 5

        img_width = len(data[0]) * cell_width if data else 800
        img_height = len(data) * cell_height if data else 600

        img = PILImage.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)

        # 셀 그리기
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                x = j * cell_width
                y = i * cell_height

                # 셀 테두리
                draw.rectangle([x, y, x + cell_width, y + cell_height], outline='gray')

                # 텍스트 (잘라내기)
                text = str(value)[:12]
                draw.text((x + padding, y + padding), text, fill='black')

        # 저장
        if output_path is None:
            fd, output_path = tempfile.mkstemp(suffix='.png')

        img.save(output_path)
        return output_path

    def extract_headers(self, file_path: str, header_rows: int = 3) -> List[List[str]]:
        """헤더 행 추출"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        headers = []
        for row_num in range(1, header_rows + 1):
            row_data = []
            for cell in sheet[row_num]:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            headers.append(row_data)

        return headers

    def extract_data_types(self, file_path: str) -> List[List[str]]:
        """셀 데이터 타입 패턴 추출"""
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        type_pattern = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10)):
            row_types = []
            for cell in row:
                if cell.value is None:
                    row_types.append("empty")
                elif isinstance(cell.value, (int, float)):
                    row_types.append("number")
                else:
                    row_types.append("text")
            type_pattern.append(row_types)

        return type_pattern

    def parse_with_mapping(
        self,
        file_path: str,
        mapping: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """매핑 정보를 사용하여 Excel 데이터 파싱"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        model_col = mapping.get("model_column", "A")
        model_start_row = mapping.get("model_start_row", 3)
        date_row = mapping.get("date_row", 1)
        date_start_col = mapping.get("date_start_column", "B")
        skip_rows = mapping.get("skip_rows", [])
        skip_columns = mapping.get("skip_columns", [])

        # 날짜/주차 헤더 읽기
        dates = []
        col_idx = openpyxl.utils.column_index_from_string(date_start_col)
        while col_idx <= sheet.max_column:
            date_value = sheet.cell(row=date_row, column=col_idx).value
            if date_value and str(date_value) not in skip_columns:
                dates.append({"col": col_idx, "value": str(date_value)})
            col_idx += 1

        # 데이터 추출
        results = []
        model_col_idx = openpyxl.utils.column_index_from_string(model_col)

        for row_num in range(model_start_row, sheet.max_row + 1):
            if row_num in skip_rows:
                continue

            model = sheet.cell(row=row_num, column=model_col_idx).value
            if not model:
                continue

            for date_info in dates:
                quantity = sheet.cell(row=row_num, column=date_info["col"]).value
                if quantity and isinstance(quantity, (int, float)):
                    results.append({
                        "model": str(model),
                        "period": date_info["value"],
                        "quantity": int(quantity)
                    })

        return results

    def parse_cnc_forecast(self, file_path: str) -> Dict[str, Any]:
        """
        CNC Forecast Excel 파일 파싱 (고정 형식)

        실제 형식 (스크린샷 기반):
        - Row 3: 첫 번째 "⊙ Forecast CNC" (요약 섹션)
        - Row 11: 두 번째 "⊙ Forecast CNC" (실제 데이터 섹션 시작)
        - Row 12: "Model"(F열), "Process"(G열), "Vendor"(H열) 헤더
        - Row 13: 날짜들 (I열부터: 11/17, 11/18, ...)
        - Row 14+: 실제 데이터 (M1, M3, B7 Main mmW 등)

        주의: A-E열은 숨겨진 열일 수 있음. Model은 F열(6)에 있음
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        print(f"[DEBUG] Parsing CNC Forecast: max_row={sheet.max_row}, max_col={sheet.max_column}")

        # 1. 데이터 섹션 시작 찾기 - 두 번째 "Forecast CNC" 찾기
        forecast_rows = []

        for row in range(1, min(30, sheet.max_row + 1)):
            for col in range(1, min(20, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row, column=col).value
                if cell_val:
                    cell_str = str(cell_val).lower().strip()
                    if 'forecast' in cell_str and 'cnc' in cell_str:
                        forecast_rows.append(row)
                        print(f"[DEBUG] Found 'Forecast CNC' at row {row}, col {col}")
                        break

        # 두 번째 "Forecast CNC" 섹션 사용
        data_section_start = 1
        if len(forecast_rows) >= 2:
            data_section_start = forecast_rows[1]
            print(f"[DEBUG] Using second Forecast CNC section starting at row {data_section_start}")
        elif len(forecast_rows) == 1:
            data_section_start = forecast_rows[0]
            print(f"[DEBUG] Using first Forecast CNC section starting at row {data_section_start}")

        # 2. "Model" 헤더 찾기 (모든 열에서 검색)
        model_header_row = 0
        model_col = 1  # 기본값

        # 디버그: data_section 이후 행들의 내용 출력
        print(f"[DEBUG] Searching for 'Model' header from row {data_section_start}")
        for row in range(data_section_start, min(data_section_start + 5, sheet.max_row + 1)):
            row_vals = []
            for col in range(1, min(15, sheet.max_column + 1)):
                val = sheet.cell(row=row, column=col).value
                row_vals.append(f"{col}:{str(val)[:15]}" if val else "")
            print(f"[DEBUG] Row {row}: {[v for v in row_vals if v]}")

        for row in range(data_section_start, min(data_section_start + 10, sheet.max_row + 1)):
            for col in range(1, min(20, sheet.max_column + 1)):
                cell_val = sheet.cell(row=row, column=col).value
                if cell_val and 'model' in str(cell_val).lower():
                    model_header_row = row
                    model_col = col
                    print(f"[DEBUG] Found 'Model' header at row {row}, col {col}: '{cell_val}'")
                    break
            if model_header_row > 0:
                break

        if model_header_row == 0:
            print(f"[DEBUG] WARNING: 'Model' header not found! Using fallback.")

        # Process, Vendor 열 위치 (Model 다음 열들)
        process_col = model_col + 1
        vendor_col = model_col + 2
        data_start_col = model_col + 3  # 데이터 시작 열

        print(f"[DEBUG] Column mapping: Model={model_col}, Process={process_col}, Vendor={vendor_col}, Data starts at col {data_start_col}")

        # 3. 날짜 행 찾기 - Model 헤더 다음 행에서 날짜 패턴 검색
        date_columns = {}  # {col_idx: date_string}
        current_year = datetime.now().year
        date_row = model_header_row + 1

        # 날짜 행 후보들 검색 (Model 헤더 이후 3개 행까지)
        for check_row in range(model_header_row + 1, min(model_header_row + 4, sheet.max_row + 1)):
            found_dates = 0
            for col in range(data_start_col, sheet.max_column + 1):
                cell_val = sheet.cell(row=check_row, column=col).value
                if cell_val:
                    # datetime 객체인 경우
                    if isinstance(cell_val, datetime):
                        found_dates += 1
                    else:
                        # 문자열인 경우 MM/DD 패턴 체크
                        cell_str = str(cell_val).strip()
                        match = re.match(r'(\d{1,2})/(\d{1,2})', cell_str)
                        if match:
                            found_dates += 1
            if found_dates >= 3:
                date_row = check_row
                print(f"[DEBUG] Found date row at row {check_row} with {found_dates} dates")
                break

        # 날짜 컬럼 매핑 (찾은 날짜 행에서)
        for col in range(data_start_col, sheet.max_column + 1):
            cell_val = sheet.cell(row=date_row, column=col).value
            if cell_val:
                # datetime 객체인 경우 직접 변환
                if isinstance(cell_val, datetime):
                    date_str = cell_val.strftime("%Y-%m-%d")
                    date_columns[col] = date_str
                    print(f"[DEBUG] Col {col}: datetime -> {date_str}")
                else:
                    # 문자열인 경우 MM/DD 패턴 체크
                    cell_str = str(cell_val).strip()
                    match = re.match(r'(\d{1,2})/(\d{1,2})', cell_str)
                    if match:
                        month = int(match.group(1))
                        day = int(match.group(2))
                        year = current_year if month >= datetime.now().month - 1 else current_year + 1
                        date_str = f"{year}-{month:02d}-{day:02d}"
                        date_columns[col] = date_str

        print(f"[DEBUG] Found {len(date_columns)} date columns from row {date_row}")

        # 4. 데이터 시작 행 찾기 - 날짜 행 다음 행부터
        data_start_row = date_row + 1
        skip_keywords = ['total', '합계', 'sum', 'ag tech', 'agtech']

        # 5. 데이터 파싱
        results = []
        current_model = None

        for row_num in range(data_start_row, sheet.max_row + 1):
            # 모델명 (병합셀 처리 - 비어있으면 이전 값 유지)
            model_cell = sheet.cell(row=row_num, column=model_col).value
            if model_cell:
                model_str = str(model_cell).strip()
                if not any(kw in model_str.lower() for kw in skip_keywords):
                    current_model = model_str

            if not current_model:
                continue

            # 공정 (Process)
            process_cell = sheet.cell(row=row_num, column=process_col).value
            process = str(process_cell).strip() if process_cell else ""

            # 빈 행 스킵
            if not process:
                continue

            # 각 날짜 컬럼의 수량 추출
            for col_idx, date_str in date_columns.items():
                quantity_cell = sheet.cell(row=row_num, column=col_idx).value

                if quantity_cell is not None:
                    try:
                        if isinstance(quantity_cell, str):
                            quantity_cell = quantity_cell.replace(',', '').strip()
                            if quantity_cell == '-' or quantity_cell == '':
                                continue
                        quantity = int(float(quantity_cell))

                        if quantity > 0:
                            results.append({
                                "model": current_model,
                                "process": process,
                                "period": date_str,
                                "quantity": quantity
                            })
                    except (ValueError, TypeError):
                        continue

        print(f"[DEBUG] Parsed {len(results)} data items")

        return {
            "success": True,
            "data": results,
            "confidence": 1.0,
            "notes": f"CNC Forecast 템플릿 파싱 완료 ({len(results)}건)",
            "template_matched": True,
            "template_name": "CNC_FORECAST_STANDARD"
        }

    def _week_day_to_date(self, week_num: int, day_offset: int) -> str:
        """
        ISO 주차 번호와 요일 오프셋으로 실제 날짜 계산

        Args:
            week_num: ISO 주차 번호 (1-52)
            day_offset: 요일 오프셋 (0=월, 1=화, ..., 6=일)

        Returns:
            날짜 문자열 (YYYY-MM-DD)
        """
        # 현재 연도 기준
        current_year = datetime.now().year

        # Week 번호가 현재 주보다 작으면 내년으로 간주
        current_week = datetime.now().isocalendar()[1]
        year = current_year if week_num >= current_week - 4 else current_year + 1

        # ISO 주차의 첫 번째 월요일 계산
        jan4 = datetime(year, 1, 4)
        start_of_week1 = jan4 - timedelta(days=jan4.weekday())

        # 해당 주차의 시작일 (월요일)
        week_start = start_of_week1 + timedelta(weeks=week_num - 1)

        # 요일 오프셋 적용
        target_date = week_start + timedelta(days=day_offset)

        return target_date.strftime("%Y-%m-%d")

    def is_cnc_forecast_format(self, file_path: str) -> bool:
        """
        CNC Forecast 형식인지 확인
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            # 디버그: 첫 3행 내용 출력
            print(f"[DEBUG] Sheet: {sheet.title}, max_row={sheet.max_row}, max_col={sheet.max_column}")
            for row in range(1, min(4, sheet.max_row + 1)):
                row_vals = []
                for col in range(1, min(15, sheet.max_column + 1)):
                    val = sheet.cell(row=row, column=col).value
                    row_vals.append(str(val)[:20] if val else "")
                print(f"[DEBUG] Row {row}: {row_vals}")

            # 체크 1: 처음 5행 내에서 "Forecast" 또는 "CNC" 키워드 찾기
            found_forecast = False
            found_week = False

            for row in range(1, min(6, sheet.max_row + 1)):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell_val = sheet.cell(row=row, column=col).value
                    if cell_val:
                        cell_str = str(cell_val).lower()
                        if 'forecast' in cell_str or 'cnc' in cell_str:
                            found_forecast = True
                            print(f"[DEBUG] Found 'forecast/cnc' at row={row}, col={col}: {cell_val}")
                        if 'week' in cell_str:
                            found_week = True
                            print(f"[DEBUG] Found 'week' at row={row}, col={col}: {cell_val}")

            result = found_forecast and found_week
            print(f"[DEBUG] is_cnc_forecast_format result: {result} (forecast={found_forecast}, week={found_week})")
            return result
        except Exception as e:
            print(f"[DEBUG] is_cnc_forecast_format error: {e}")
            return False


excel_service = ExcelService()
